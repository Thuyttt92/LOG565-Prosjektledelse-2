"""Bygger Kravspesifikasjon - Nye Hædda barneskole.xlsx med 50+ krav for A-nivå."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

PATH = r"C:\Users\Gusta\OneDrive\Documents\Privat\Studier\Bachelore i Logistikk\LOG565 - Prosjektledelse\02 - Planlegging\Kravspesifikasjon - Nye Hædda barneskole.xlsx"

# (ID, Kategori, Beskrivelse, Prioritet, WBS Ref, Ansvarlig disiplin, Verifikasjonsmetode, Kilde)
krav = [
    # === FUNKSJONELT ===
    ("F-001", "Funksjonelt", "Skolen skal dimensjoneres for 600 elever fordelt på trinn 1–10.", "Skal", "4.0", "Arkitekt", "Romprogram + tegningsgranskning", "Prosjektforslag"),
    ("F-002", "Funksjonelt", "Maksimalt 20 elever per klasserom.", "Skal", "4.2, 4.3, 4.4", "Arkitekt", "Tegningskontroll", "Pedagogisk standard"),
    ("F-003", "Funksjonelt", "Klasserom skal ha minimum 2,5 m² programareal per elev.", "Skal", "4.2, 4.3, 4.4", "Arkitekt", "Arealmåling i tegning", "Veileder skoleanlegg"),
    ("F-004", "Funksjonelt", "Skolebygget skal soneinndeles i småtrinn (1–4), mellomtrinn (5–7) og ungdomstrinn (8–10).", "Skal", "4.2, 4.3, 4.4", "Arkitekt", "Romprogram", "Konseptløsning"),
    ("F-005", "Funksjonelt", "Bygget skal romme bibliotek og auditorium med plass til minst 120 personer.", "Skal", "4.3", "Arkitekt", "Tegning + møblering", "Konseptløsning"),
    ("F-006", "Funksjonelt", "Spesialrom skal etableres for naturfag, kunst & håndverk, musikk og mat & helse.", "Skal", "4.4", "Arkitekt", "Romprogram", "Konseptløsning"),
    ("F-007", "Funksjonelt", "Gymsalen skal ha full takhøyde (≥ 7 m) og målene 24×16 m.", "Skal", "4.5", "Arkitekt", "Tegning + befaring", "Idrettsfunksjonell standard"),
    ("F-008", "Funksjonelt", "Kantinen skal kunne servere 600 elever fordelt på to spisesesjoner.", "Skal", "4.3", "Arkitekt", "Kapasitetsberegning", "Brukermøte"),
    ("F-009", "Funksjonelt", "Skolen skal ha 70 arbeidsplasser for fagansatte med tilstrekkelig hvile- og pauserom.", "Skal", "4.6", "Arkitekt", "Romprogram + arbeidsmiljøvurdering", "Arbeidsmiljøloven"),
    ("F-010", "Funksjonelt", "Driftssone med vaktmesterkontor, lager og verksted skal etableres separat.", "Skal", "4.6", "Arkitekt", "Tegning", "FDV-krav"),
    ("F-011", "Funksjonelt", "Helseavdeling med skolelegerom og helsesøsterkontor skal etableres på 1. etasje.", "Skal", "4.2", "Arkitekt", "Tegning", "Brukermøte"),
    ("F-012", "Funksjonelt", "Garderober og dusjanlegg skal være tilstrekkelige for samtidig bruk av 60 elever.", "Skal", "4.5", "Arkitekt", "Kapasitetsberegning", "Brukermøte"),
    ("F-013", "Funksjonelt", "SFO-areal skal etableres med separat inngang og direkte tilgang til uteområde.", "Skal", "4.2", "Arkitekt", "Tegning", "Brukermøte"),
    ("F-014", "Funksjonelt", "Fellesarealer skal være tilstrekkelig fleksible til å kunne brukes som scene/samlingsrom.", "Bør", "4.3", "Arkitekt", "Tegning + møbleringsplan", "Konseptløsning"),
    ("F-015", "Funksjonelt", "Resepsjon/administrasjon skal plasseres tydelig synlig fra hovedinngang.", "Skal", "4.6", "Arkitekt", "Tegning", "Sikkerhet/funksjonell standard"),

    # === TEKNISK ===
    ("T-001", "Teknisk", "Bygget skal oppfylle TEK17 og gjeldende energi- og branntekniske krav.", "Skal", "2.0, 5.0", "RIB", "Sjekkliste mot TEK17", "Lovkrav"),
    ("T-002", "Teknisk", "Balansert ventilasjon med varmegjenvinning (≥ 80 % virkningsgrad) skal installeres.", "Skal", "5.1", "VVS", "Funksjonstest + dokumentasjon", "TEK17 §13"),
    ("T-003", "Teknisk", "Innetemperatur skal kunne holdes mellom 20 og 24 °C i bruksfasen.", "Skal", "5.1", "VVS", "Måling i prøvedrift", "TEK17 §13"),
    ("T-004", "Teknisk", "Vannbåren varme med fjernvarmetilkobling skal være primær oppvarmingskilde.", "Skal", "5.1", "VVS", "Tegninger + tilkoblingsdokumentasjon", "Miljøstrategi"),
    ("T-005", "Teknisk", "Trådløst nett skal gi ≥ 100 Mbps i alle rom og dekning ved alle elev- og personalplasser.", "Skal", "5.5", "IKT", "Dekningsmåling", "Brukerkrav"),
    ("T-006", "Teknisk", "Sentralt driftsovervåkningsanlegg (SD-anlegg) skal styre lys, varme, ventilasjon og adgangskontroll.", "Skal", "5.4", "Automasjon", "Funksjonstest", "FDV-strategi"),
    ("T-007", "Teknisk", "Heis mellom alle etasjer skal ha minimum løftekapasitet 1000 kg og rom for båre.", "Skal", "5.3", "Heisleverandør", "Sertifisering + funksjonstest", "TEK17 §12-6"),
    ("T-008", "Teknisk", "LED-belysning med dagslys- og tilstedeværelsesstyring skal benyttes i alle rom.", "Skal", "5.2", "Elektro", "Tegning + funksjonstest", "Energistrategi"),
    ("T-009", "Teknisk", "Reservekraft (UPS) skal sikre IKT-rom og nødlys i minst 2 timer ved strømbrudd.", "Skal", "5.2", "Elektro", "Funksjonstest", "Sikkerhetsstrategi"),
    ("T-010", "Teknisk", "AV-utstyr (skjerm + lyd) skal installeres i alle klasserom og auditorium.", "Skal", "7.3", "IKT", "Funksjonstest", "Pedagogisk strategi"),
    ("T-011", "Teknisk", "Dataservere og IKT-skap skal plasseres i kjølt og adgangskontrollert serverrom.", "Skal", "5.5", "IKT", "Befaring + dokumentasjon", "ISO 27001"),

    # === MILJØ ===
    ("M-001", "Miljø", "Prosjektet skal sertifiseres etter BREEAM-NOR med klasse Very Good eller bedre.", "Skal", "2.0", "Miljøkoordinator", "Sertifikat", "Konseptløsning"),
    ("M-002", "Miljø", "Sedum-tak skal etableres som overvannshåndtering på minst 50 % av takflaten.", "Skal", "4.7", "Arkitekt", "Tegning + befaring", "Konseptløsning"),
    ("M-003", "Miljø", "Det skal være ladeplasser for minst 20 elbiler og 4 elsykler.", "Skal", "6.3", "Elektro", "Tegning + leveransedokumentasjon", "Konseptløsning"),
    ("M-004", "Miljø", "Minimum 50 % av byggematerialer skal være gjenvinnbare eller fra fornybare kilder.", "Bør", "4.0, 5.0", "Miljøkoordinator", "Materialregister", "BREEAM"),
    ("M-005", "Miljø", "Energiforbruk i drift skal ikke overskride 75 kWh/m²/år.", "Skal", "5.0", "RIE/RIV", "Energiberegning", "Energistrategi"),
    ("M-006", "Miljø", "Avfall i byggefase skal sorteres og minimum 85 % skal materialgjenvinnes.", "Skal", "4.0", "Miljøkoordinator", "Avfallsplan + sluttrapport", "BREEAM"),

    # === SIKKERHET ===
    ("S-001", "Sikkerhet", "Sprinkleranlegg skal installeres i hele bygget.", "Skal", "5.1", "VVS/Brann", "Sertifisering + funksjonstest", "Brannkonsept"),
    ("S-002", "Sikkerhet", "Rømningsveier skal dimensjoneres iht. brannkonsept og merkes tydelig.", "Skal", "4.0", "Brannrådgiver", "Brannteknisk kontroll", "TEK17"),
    ("S-003", "Sikkerhet", "Adgangskontroll med loggføring skal etableres på alle yttredører.", "Skal", "5.5", "IKT/Elektro", "Funksjonstest", "Sikkerhetsstrategi"),
    ("S-004", "Sikkerhet", "Kameraovervåkning skal etableres i fellesarealer ute og ved hovedinnganger.", "Skal", "5.5", "IKT/Elektro", "Funksjonstest", "Sikkerhetsstrategi"),
    ("S-005", "Sikkerhet", "Innbruddsalarm skal dekke alle rom med IT-utstyr og verdier.", "Skal", "5.5", "IKT", "Funksjonstest + sertifisering", "Forsikringskrav"),
    ("S-006", "Sikkerhet", "Skolens hovedinngang skal kunne låses sentralt fra resepsjon (lockdown).", "Skal", "5.5", "IKT", "Funksjonstest", "Sikkerhetsstrategi"),

    # === UTEOMRÅDE ===
    ("U-001", "Uteområde", "Lekearealer skal etableres med fallunderlag i hht. NS-EN 1177 for alle alderstrinn.", "Skal", "6.1", "Landskap", "Befaring + sertifikat", "NS-EN 1177"),
    ("U-002", "Uteområde", "Det skal være ballbinge med kunstgress (min 20×40 m).", "Skal", "6.2", "Landskap", "Tegning + befaring", "Konseptløsning"),
    ("U-003", "Uteområde", "Sykkelparkering under tak skal romme minimum 100 sykler.", "Skal", "6.3", "Landskap", "Befaring", "Konseptløsning"),
    ("U-004", "Uteområde", "Hele uteområdet skal ha belysning styrt av tilstedeværelses- og dagslyssensor.", "Skal", "6.2", "Elektro", "Funksjonstest", "Sikkerhetsstrategi"),
    ("U-005", "Uteområde", "Skolegården skal skjermes mot vei med støyskjerm eller beplantning.", "Skal", "6.4", "Landskap", "Befaring + støymåling", "Reguleringsplan"),
    ("U-006", "Uteområde", "Det skal være tydelig adskilte soner for små-, mellom- og ungdomstrinn ute.", "Skal", "6.1", "Landskap", "Tegning + befaring", "Pedagogisk strategi"),
    ("U-007", "Uteområde", "Adkomst for bringe-/hentetrafikk skal være adskilt fra elevtrafikk.", "Skal", "6.3", "Landskap", "Tegning + befaring", "Sikkerhetsstrategi"),

    # === AKUSTIKK ===
    ("A-001", "Akustikk", "Klasserom skal ha etterklangstid ≤ 0,5 sekund (NS 8175 klasse C).", "Skal", "4.2, 4.3, 4.4", "RIA", "Akustisk måling", "NS 8175"),
    ("A-002", "Akustikk", "Gymsal skal oppfylle NS 8175 klasse C med tilpassede absorbenter.", "Skal", "4.5", "RIA", "Akustisk måling", "NS 8175"),
    ("A-003", "Akustikk", "Fellesarealer skal ha lydabsorbenter som demper støy fra elever (Lp ≤ 55 dB ved normal aktivitet).", "Skal", "4.3", "RIA", "Akustisk måling", "NS 8175"),

    # === UNIVERSELL UTFORMING ===
    ("UU-001", "Universell utforming", "Hele skolen skal være universelt utformet iht. TEK17 §12.", "Skal", "4.0", "Arkitekt", "UU-sjekkliste", "TEK17 §12"),
    ("UU-002", "Universell utforming", "Alle etasjer skal være tilgjengelige med rullestolheis og rampe.", "Skal", "5.3", "Arkitekt/Heis", "Befaring", "TEK17"),
    ("UU-003", "Universell utforming", "Ledelinjer og kontrastmarkering skal etableres ved trapper og kjernepunkter.", "Skal", "4.0", "Arkitekt", "Befaring", "TEK17 §12"),
    ("UU-004", "Universell utforming", "Teleslynge skal installeres i auditorium og resepsjon.", "Skal", "5.5", "IKT", "Funksjonstest", "Likestillingsloven"),

    # === KVALITET ===
    ("K-001", "Kvalitet", "Bygget skal overleveres med null kritiske mangler ved BP3-ferdigbefaring.", "Skal", "8.2", "PL", "Befaringsrapport", "Kontraktsstandard"),
    ("K-002", "Kvalitet", "Fullstendig FDV-dokumentasjon skal leveres digitalt før overtakelse.", "Skal", "8.3", "PL/FDV", "Dokumentkontroll", "FDV-strategi"),
    ("K-003", "Kvalitet", "Garantitid på bygningstekniske og tekniske leveranser skal være minimum 5 år.", "Skal", "8.0", "PL", "Kontraktsdokumentasjon", "NS 8407"),

    # === DRIFT / FDV ===
    ("D-001", "Drift/FDV", "Materialer i fellesarealer skal være slitesterke og lette å renholde (klasse 33+).", "Skal", "4.0", "Arkitekt", "Materialregister", "FDV-strategi"),
    ("D-002", "Drift/FDV", "Energi- og vannforbruk skal kunne avleses og rapporteres pr. sone via SD-anlegg.", "Skal", "5.4", "Automasjon", "Funksjonstest", "FDV-strategi"),
    ("D-003", "Drift/FDV", "Reservedeler for kritiske tekniske komponenter skal lagerføres på skolen.", "Bør", "5.0", "FDV-leder", "Lagersjekk", "FDV-strategi"),
    ("D-004", "Drift/FDV", "Brukeropplæring av driftspersonell skal gjennomføres før overtakelse.", "Skal", "8.4", "Driftssjef", "Opplæringsbevis", "Kontraktskrav"),
]

print(f"Antall krav: {len(krav)}")
assert len(krav) >= 45, "For få krav!"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Kravspesifikasjon"

# Topptekst
ws["A1"] = "KRAVSPESIFIKASJON — Nye Hædda barneskole"
ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.merge_cells("A1:H1")
ws.row_dimensions[1].height = 28

ws["A2"] = "Versjon 1.0 — 04.05.2026 — Sendt til Bård for tids- og kostnadsestimat"
ws["A2"].font = Font(italic=True, size=10)
ws["A2"].alignment = Alignment(horizontal="center")
ws.merge_cells("A2:H2")

# Header
headers = ["ID", "Kategori", "Beskrivelse (Skal/Bør-krav)", "Prioritet", "WBS Ref", "Ansvarlig disiplin", "Verifikasjonsmetode", "Kilde"]
header_row = 4
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=header_row, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2E75B6")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Farger pr kategori
cat_colors = {
    "Funksjonelt": "DEEBF7",
    "Teknisk": "E2EFDA",
    "Miljø": "C6E0B4",
    "Sikkerhet": "FCE4D6",
    "Uteområde": "FFF2CC",
    "Akustikk": "EDEDED",
    "Universell utforming": "D9E1F2",
    "Kvalitet": "F8CBAD",
    "Drift/FDV": "FFD966",
}

thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for i, (kid, kat, beskr, prio, wbs, ansv, verif, kilde) in enumerate(krav):
    r = header_row + 1 + i
    row_vals = [kid, kat, beskr, prio, wbs, ansv, verif, kilde]
    fill = PatternFill("solid", fgColor=cat_colors.get(kat, "FFFFFF"))
    for c, v in enumerate(row_vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if c == 1:
            cell.font = Font(bold=True)

# Kolonnebredder
widths = [10, 22, 70, 12, 14, 22, 30, 22]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Frys topp
ws.freeze_panes = "A5"

# Statistikk-ark
ws2 = wb.create_sheet("Statistikk")
ws2["A1"] = "Statistikk for kravspesifikasjon"
ws2["A1"].font = Font(bold=True, size=14)
ws2["A3"] = "Kategori"
ws2["B3"] = "Antall krav"
ws2["A3"].font = Font(bold=True)
ws2["B3"].font = Font(bold=True)
from collections import Counter
cnt = Counter(k[1] for k in krav)
r = 4
for k, v in sorted(cnt.items()):
    ws2.cell(row=r, column=1, value=k)
    ws2.cell(row=r, column=2, value=v)
    r += 1
ws2.cell(row=r, column=1, value="TOTALT").font = Font(bold=True)
ws2.cell(row=r, column=2, value=len(krav)).font = Font(bold=True)
ws2["A1"].alignment = Alignment(horizontal="left")
ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 15

# Instruksjons-ark for Bård
ws3 = wb.create_sheet("Til Bård")
ws3["A1"] = "Til Bård — leseveiledning"
ws3["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws3["A1"].fill = PatternFill("solid", fgColor="1F4E79")
ws3.merge_cells("A1:F1")
notes = [
    "",
    "Denne kravspesifikasjonen er bygget som grunnlag for at du skal kunne sette tids- og kostnadsestimat",
    "i WBS-filen («WBS - Nye Hædda barneskole.xlsx»). Hver leveranse i WBS er forsøkt knyttet opp mot",
    "ett eller flere krav her via 'WBS Ref'-kolonnen.",
    "",
    "Strukturen følger PMI-malen og er kategorisert slik:",
    "  • Funksjonelt — krav fra prosjektforslag/konseptløsning og brukermøter",
    "  • Teknisk — TEK17, ventilasjon, IKT, automasjon",
    "  • Miljø — BREEAM, energi, materialer, avfall",
    "  • Sikkerhet — brann, adgang, kamera, alarm",
    "  • Uteområde — lek, sport, infrastruktur, sikkerhet",
    "  • Akustikk — NS 8175 klasse C",
    "  • Universell utforming — TEK17 §12",
    "  • Kvalitet — overtakelse, FDV, garanti",
    "  • Drift/FDV — slitestyrke, energi, opplæring",
    "",
    f"Totalt antall krav: {len(krav)} (mål var ≥ 40 for A-nivå).",
    "Ved spørsmål, ta kontakt på Teams.",
    "",
    "Mvh",
    "Studentgruppe LOG565",
]
for i, t in enumerate(notes, start=2):
    ws3.cell(row=i, column=1, value=t)
ws3.column_dimensions["A"].width = 110

wb.save(PATH)
print(f"Lagret: {PATH}")
