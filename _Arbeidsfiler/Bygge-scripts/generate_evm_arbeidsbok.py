"""
Genererer EVM-arbeidsboken for Nye Hædda Barneskole — 16 måneder fase 3.

Ark:
  - Innstillinger: prosjektrammer (BAC, risikoreserve, tidsbuffer)
  - Inngang: månedlig PV/AC/% complete (kalkulert fra master data)
  - Dashboard: KPI-er per måned + S-kurve datatabell
  - Risks_Log: risikoregister + hendelser
  - S_Curve_Data: rådata for S-kurven
  - Bruksanvisning

Replicerer EVM-arbeidsbok-mal.xlsx-strukturen, men utvidet til 16 mnd og
befolket med faktiske tall fra master_data.py.
"""
from __future__ import annotations
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from log565_master_data import (
    BAC, RISIKORESERVE, TIDSBUFFER_UKER, PROSJEKTSTART, SLUTTDATO,
    MÅNEDER, PAKKER, HENDELSER,
    AC_KUM, RISIKORESERVE_BRUKT_KUM, TIDSBUFFER_BRUKT_KUM_UKER,
    beregn_evm, beregn_pv, beregn_ev, alle_måneder,
)

sys.path.insert(0, str(Path(__file__).resolve().parent))
from paths import ROOT, GJENNOMFORING
OUT = GJENNOMFORING / "EVM-arbeidsbok - Nye Hædda barneskole.xlsx"
BAC_KR = BAC * 1_000_000  # MNOK → kr
RISIKO_KR = RISIKORESERVE * 1_000_000

# =============================================================================
# Styling helpers
# =============================================================================
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SUBHDR_FILL = PatternFill("solid", fgColor="BDD7EE")
SUBHDR_FONT = Font(name="Calibri", size=10, bold=True, color="1F4E78")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")  # GUL = manuell input
CALC_FILL = PatternFill("solid", fgColor="E2EFDA")  # GRØNN = beregnet
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")  # ORANSJE = avvik
THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def style_header_row(ws, row, end_col):
    for c in range(1, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


# =============================================================================
# Build workbook
# =============================================================================
def build_innstillinger(wb):
    ws = wb.active
    ws.title = "Innstillinger"
    ws["A1"] = "EVM-arbeidsbok — Nye Hædda Barneskole"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
    ws.merge_cells("A1:C1")

    headers = ["Parameter", "Verdi", "Kommentar"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 3)

    rows = [
        ("Prosjektnavn", "Nye Hædda Barneskole", "Fast"),
        ("Prosjektleder (gruppe)", "irgesundinger", "Gruppe 4.5"),
        ("Saksnummer (kommunestyret)", "NHB-IRGESUND", "Crashing + rammeutvidelse"),
        ("BAC (Budget at Completion)", BAC_KR, f"= {BAC:.0f} MNOK — Baseline 1 etter crashing av 4.1 Råbygg"),
        ("Godkjent risikoreserve", RISIKO_KR, f"= {RISIKORESERVE:.0f} MNOK (separat fra BAC)"),
        ("Godkjent tidsbuffer", TIDSBUFFER_UKER, "Uker ut over sluttdato før forsinkelse er reell"),
        ("Prosjektstart", PROSJEKTSTART, "Fra Baseline 1 (MS Project)"),
        ("Prosjektslutt (vedtatt)", SLUTTDATO, "Fra Baseline 1 (MS Project)"),
        ("Antall måneder (gjennomføring)", 16, "Feb 2025 → Mai 2026"),
        ("Antall arbeidspakker (leaves)", len(PAKKER), "32 leaves i 4-nivå WBS"),
        ("Valuta", "NOK (kr)", "Alle beløp vises i kr"),
        ("Versjon", "1.0 — sluttall etter mnd 16", "Generert fra master_data.py"),
    ]
    for i, (k, v, c) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v).fill = CALC_FILL
        ws.cell(row=i, column=3, value=c)
        ws.cell(row=i, column=2).number_format = '#,##0'
        if k == "Godkjent tidsbuffer":
            ws.cell(row=i, column=2).number_format = "0"
    set_col_widths(ws, [32, 30, 60])


def build_inngang(wb):
    """Månedlig input: PV planlagt, AC faktisk, % complete kum."""
    ws = wb.create_sheet("Inngang")
    ws["A1"] = "Månedlig EVM-inngang — PV planlagt, AC faktisk, % complete"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
    ws.merge_cells("A1:F1")
    ws["A2"] = "GRØNN = avledet fra master_data.py (referater + Baseline 1). Endre i master_data.py, ikke her direkte."
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:F2")

    headers = ["Måned (nr)", "Periode", "PV planlagt (kr)", "AC faktisk (kr)",
               "% complete kum.", "Risikoreserve brukt kum. (kr)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, len(headers))

    forrige_pv = 0
    for m in range(1, 17):
        navn, statusdato = MÅNEDER[m]
        evm = beregn_evm(m)
        pv_periode = (evm.pv_kum - forrige_pv) * 1_000_000
        forrige_pv = evm.pv_kum
        ac_kum = AC_KUM[m] * 1_000_000
        pct = evm.pct_complete / 100
        rr_kum = RISIKORESERVE_BRUKT_KUM[m] * 1_000_000

        r = 4 + m
        ws.cell(row=r, column=1, value=m).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=navn)
        ws.cell(row=r, column=3, value=pv_periode).number_format = '#,##0'
        ws.cell(row=r, column=4, value=ac_kum).number_format = '#,##0'
        ws.cell(row=r, column=5, value=pct).number_format = '0.0%'
        ws.cell(row=r, column=6, value=rr_kum).number_format = '#,##0'
        for c in range(3, 7):
            ws.cell(row=r, column=c).fill = CALC_FILL
            ws.cell(row=r, column=c).border = BORDER

    # Sumrad
    r = 21
    ws.cell(row=r, column=1, value="SUM").font = Font(bold=True)
    ws.cell(row=r, column=2, value="(PV → BAC, AC → faktisk total)")
    ws.cell(row=r, column=3, value=f"=SUM(C5:C{r-1})").number_format = '#,##0'
    ws.cell(row=r, column=4, value=f"=MAX(D5:D{r-1})").number_format = '#,##0'
    for c in range(3, 5):
        ws.cell(row=r, column=c).fill = SUBHDR_FILL
        ws.cell(row=r, column=c).font = Font(bold=True)

    set_col_widths(ws, [12, 18, 20, 20, 16, 28])


def build_dashboard(wb):
    """KPI-er per måned + S-kurve-datatabell."""
    ws = wb.create_sheet("Dashboard")
    ws["A1"] = "Dashboard — EVM-KPI-er og S-kurve"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
    ws.merge_cells("A1:J1")

    # KPI-strip for siste måned
    ws["A3"] = "KPI-sammendrag (sluttall mnd 16):"
    ws["A3"].font = Font(bold=True, size=11)
    headers = ["BAC", "PV kum", "EV kum", "AC kum", "CPI", "SPI", "EAC", "VAC", "% Complete"]
    for i, h in enumerate(headers, 2):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, 10)

    siste = beregn_evm(16)
    vals = [BAC_KR, siste.pv_kum*1e6, siste.ev_kum*1e6, siste.ac_kum*1e6,
            siste.cpi, siste.spi, siste.eac*1e6, siste.vac*1e6, siste.pct_complete/100]
    formats = ['#,##0', '#,##0', '#,##0', '#,##0', '0.000', '0.000', '#,##0', '#,##0', '0.0%']
    for i, (v, f) in enumerate(zip(vals, formats), 2):
        cell = ws.cell(row=5, column=i, value=v)
        cell.number_format = f
        cell.fill = CALC_FILL
        cell.border = BORDER

    # Måned-for-måned EVM
    ws["A7"] = "Måned-for-måned EVM"
    ws["A7"].font = Font(bold=True, size=11)
    headers = ["Måned", "Periode", "PV kum (kr)", "EV kum (kr)", "AC kum (kr)",
               "CPI", "SPI", "EAC (kr)", "VAC (kr)", "Status"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=8, column=i, value=h)
    style_header_row(ws, 8, len(headers))

    for m in range(1, 17):
        evm = beregn_evm(m)
        navn, _ = MÅNEDER[m]
        r = 8 + m
        ws.cell(row=r, column=1, value=m).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=navn)
        ws.cell(row=r, column=3, value=evm.pv_kum*1e6).number_format = '#,##0'
        ws.cell(row=r, column=4, value=evm.ev_kum*1e6).number_format = '#,##0'
        ws.cell(row=r, column=5, value=evm.ac_kum*1e6).number_format = '#,##0'
        ws.cell(row=r, column=6, value=evm.cpi).number_format = '0.000'
        ws.cell(row=r, column=7, value=evm.spi).number_format = '0.000'
        ws.cell(row=r, column=8, value=evm.eac*1e6).number_format = '#,##0'
        ws.cell(row=r, column=9, value=evm.vac*1e6).number_format = '#,##0'
        ws.cell(row=r, column=10, value=evm.status_rag)
        for c in range(3, 10):
            ws.cell(row=r, column=c).fill = CALC_FILL
            ws.cell(row=r, column=c).border = BORDER
        # Status-fargekoding
        if evm.status_rag == "Rød":
            ws.cell(row=r, column=10).fill = PatternFill("solid", fgColor="F4B084")
        elif evm.status_rag == "Gul":
            ws.cell(row=r, column=10).fill = PatternFill("solid", fgColor="FFE699")
        else:
            ws.cell(row=r, column=10).fill = PatternFill("solid", fgColor="C6EFCE")

    set_col_widths(ws, [8, 18, 16, 16, 16, 10, 10, 16, 16, 12])


def build_risks(wb):
    """Risks_Log: hendelseslogg knyttet til risiko-ID-er fra referatene."""
    ws = wb.create_sheet("Risks_Log")
    ws["A1"] = "Risikoregister og hendelsesloggen — fase 3"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
    ws.merge_cells("A1:K1")
    ws["A2"] = ("Hendelser som har utløst respons (risikoreserve eller tidsbuffer) "
                "registrert mot risikoregisterets ID-er. CR-koblinger noteres i kolonne K.")
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:K2")

    headers = ["Risiko-ID", "Måned", "Periode", "Pakke", "Hendelse", "Kostnad (kr)",
               "Tid (uker)", "Tidsbuffer brukt?", "Beslutning", "Status", "CR-kobling"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, len(headers))

    for i, h in enumerate(HENDELSER, start=5):
        navn, _ = MÅNEDER[h.måned]
        ws.cell(row=i, column=1, value=h.risiko_id)
        ws.cell(row=i, column=2, value=h.måned).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=3, value=navn)
        ws.cell(row=i, column=4, value=h.pakke_påvirket)
        ws.cell(row=i, column=5, value=h.tittel)
        ws.cell(row=i, column=6, value=h.kostnad_mnok*1e6).number_format = '#,##0'
        ws.cell(row=i, column=7, value=h.tidsforskyvning_uker).number_format = '0.0'
        ws.cell(row=i, column=8, value="Ja" if h.tidsbuffer_brukt else "Nei (absorbert)")
        ws.cell(row=i, column=9, value=h.beslutning[:80] + "…" if len(h.beslutning) > 80 else h.beslutning)
        ws.cell(row=i, column=10, value="Lukket")
        ws.cell(row=i, column=11, value=h.cr_id or "—")
        for c in range(1, 12):
            ws.cell(row=i, column=c).border = BORDER

    # Sammendrag
    sum_r = 5 + len(HENDELSER) + 1
    ws.cell(row=sum_r, column=1, value="SUM brukt").font = Font(bold=True)
    ws.cell(row=sum_r, column=6, value=f"=SUM(F5:F{sum_r-2})").number_format = '#,##0'
    ws.cell(row=sum_r, column=7, value=f"=SUM(G5:G{sum_r-2})").number_format = '0.0'
    for c in [1, 6, 7]:
        ws.cell(row=sum_r, column=c).fill = SUBHDR_FILL
        ws.cell(row=sum_r, column=c).font = Font(bold=True)

    ws.cell(row=sum_r+1, column=1, value="Godkjent").font = Font(bold=True)
    ws.cell(row=sum_r+1, column=6, value=RISIKO_KR).number_format = '#,##0'
    ws.cell(row=sum_r+1, column=7, value=TIDSBUFFER_UKER).number_format = '0'

    ws.cell(row=sum_r+2, column=1, value="Gjenværende").font = Font(bold=True)
    ws.cell(row=sum_r+2, column=6, value=f"=F{sum_r+1}-F{sum_r}").number_format = '#,##0'
    ws.cell(row=sum_r+2, column=7, value=f"=G{sum_r+1}-G{sum_r}").number_format = '0.0'

    set_col_widths(ws, [10, 8, 16, 8, 60, 16, 10, 18, 60, 10, 12])


def build_s_curve_data(wb):
    """S_Curve_Data: rådata for S-kurven (Måned, PV/EV/AC kum)."""
    ws = wb.create_sheet("S_Curve_Data")
    ws["A1"] = "S-kurve datatabell — PV/EV/AC kum per måned"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
    ws.merge_cells("A1:G1")

    headers = ["Måned", "Periode", "PV kum (MNOK)", "EV kum (MNOK)",
               "AC kum (MNOK)", "CPI", "SPI"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(headers))

    for m in range(1, 17):
        evm = beregn_evm(m)
        navn, _ = MÅNEDER[m]
        r = 3 + m
        ws.cell(row=r, column=1, value=m).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=navn)
        ws.cell(row=r, column=3, value=evm.pv_kum).number_format = '0.0'
        ws.cell(row=r, column=4, value=evm.ev_kum).number_format = '0.0'
        ws.cell(row=r, column=5, value=evm.ac_kum).number_format = '0.0'
        ws.cell(row=r, column=6, value=evm.cpi).number_format = '0.000'
        ws.cell(row=r, column=7, value=evm.spi).number_format = '0.000'
        for c in range(3, 8):
            ws.cell(row=r, column=c).fill = CALC_FILL
            ws.cell(row=r, column=c).border = BORDER

    set_col_widths(ws, [8, 18, 16, 16, 16, 10, 10])


def build_bruksanvisning(wb):
    ws = wb.create_sheet("Bruksanvisning")
    lines = [
        ("EVM-arbeidsbok — Bruksanvisning", "header"),
        ("", ""),
        ("FORMÅL", "subhdr"),
        ("Sentral kilde for KPI-er, S-kurve og avviksanalyse for LOG565 fase 3.", ""),
        ("Arbeidsboken er generert programmatisk fra master_data.py (Arbeidsfiler/).", ""),
        ("", ""),
        ("STRUKTUR", "subhdr"),
        ("  • Innstillinger — prosjektrammer (BAC, risikoreserve, tidsbuffer)", ""),
        ("  • Inngang — månedlig PV/AC/% complete kalkulert fra referater + Baseline 1", ""),
        ("  • Dashboard — KPI-strip + måned-for-måned EVM (PV/EV/AC/CPI/SPI/EAC/VAC)", ""),
        ("  • Risks_Log — hendelseslogg knyttet til risikoregisterets ID-er", ""),
        ("  • S_Curve_Data — rådata for S-kurven (brukes til å eksportere bilde)", ""),
        ("", ""),
        ("BRUK I MÅNEDSRAPPORT", "subhdr"),
        ("  1. Hent KPI-tall fra 'Dashboard' for den aktuelle måneden", ""),
        ("  2. Hent S-kurve fra eksportert PNG (Arbeidsfiler/s_kurve_*.png)", ""),
        ("  3. Hent hendelses-/avviksdata fra 'Risks_Log'", ""),
        ("  4. Sjekk sporbarhet mot 'månedsrapporter.pdf' (Bårds rådata)", ""),
        ("", ""),
        ("KILDESPOR", "subhdr"),
        ("  • BAC og crashing-vedtak: godkjenning-av-budsjettendring.pdf (NHB-IRGESUND)", ""),
        ("  • Faktisk fremdrift og hendelser: månedsrapporter.pdf (16 teamledermøter)", ""),
        ("  • WBS og baseline-rammer: irgesundinger_19104_..._WBS_struktur-simulated.xlsx", ""),
    ]
    for i, (txt, style) in enumerate(lines, 1):
        cell = ws.cell(row=i, column=1, value=txt)
        if style == "header":
            cell.font = Font(size=14, bold=True, color="1F4E78")
        elif style == "subhdr":
            cell.font = Font(bold=True, size=11, color="1F4E78")
    set_col_widths(ws, [100])


def main():
    wb = openpyxl.Workbook()
    build_innstillinger(wb)
    build_inngang(wb)
    build_dashboard(wb)
    build_risks(wb)
    build_s_curve_data(wb)
    build_bruksanvisning(wb)
    wb.save(OUT)
    print(f"Lagret: {OUT.name}")


if __name__ == "__main__":
    main()
