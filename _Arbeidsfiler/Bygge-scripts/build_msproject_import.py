"""Lager en MS Project-vennlig Excel-fil som kan importeres direkte til MS Project
når Bårds tids- og kostnadsestimater er klare. Bruk 'File → Open' i MS Project og velg denne filen.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from build_wbs import W

PATH = r"C:\Users\Gusta\OneDrive\Documents\Privat\Studier\Bachelore i Logistikk\LOG565 - Prosjektledelse\02 - Planlegging\Gantt-import (klar for MS Project) - Nye Hædda barneskole.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Tasks"

# MS Project gjenkjenner disse navnene direkte i Import Wizard
headers = ["ID", "WBS", "Outline Level", "Name", "Duration", "Predecessors", "Cost", "Resource Names"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Mapping av WBS-ID → row id
wbs_to_id = {w[0]: idx + 1 for idx, w in enumerate(W)}

for idx, w in enumerate(W):
    wid, niv, navn, beskr, ansv, avh, krav = w
    r = idx + 2
    # Predecessors må refereres med MS Project task-ID, ikke WBS-ID
    preds = []
    if avh:
        for d in avh.split(";"):
            d = d.strip()
            if d in wbs_to_id:
                preds.append(str(wbs_to_id[d]))
    pred_str = ",".join(preds)

    ws.cell(row=r, column=1, value=idx + 1)
    ws.cell(row=r, column=2, value=wid)
    ws.cell(row=r, column=3, value=niv)
    ws.cell(row=r, column=4, value="    " * (niv - 1) + navn)
    ws.cell(row=r, column=5, value="")  # Duration - Bård fyller inn
    ws.cell(row=r, column=6, value=pred_str)
    ws.cell(row=r, column=7, value="")  # Cost - Bård fyller inn
    ws.cell(row=r, column=8, value=ansv)

widths = [6, 12, 8, 60, 12, 18, 14, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"

# Instruksjoner-ark
ws2 = wb.create_sheet("Instruksjoner")
ws2["A1"] = "Slik importerer du til MS Project"
ws2["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws2["A1"].fill = PatternFill("solid", fgColor="1F4E79")

steps = [
    "",
    "1. Når Bård har returnert tids- og kostnadsestimater, kopier dem inn i 'Tasks'-arket:",
    "   • Duration → fra kolonne H i WBS-fila Bård sender tilbake (varighet i måneder eller dager).",
    "   • Cost → fra kolonne K i WBS-fila Bård sender tilbake (kostnad i mill. kr).",
    "",
    "2. Åpne MS Project. Klikk File → Open → Browse → bla deg frem til denne Excel-fila.",
    "   Velg 'Excel Workbook' i filtypen.",
    "",
    "3. I Import Wizard:",
    "   • Velg 'New project'.",
    "   • Velg 'Use existing map'.",
    "   • Velg 'Default task information map'.",
    "   • Velg 'Tasks'-arket og bekreft kolonnemappingen.",
    "",
    "4. Etter import:",
    "   • Sett 'Project Start Date' (fra Bårds Start-felt for første aktivitet).",
    "   • Sett baseline (Project → Set Baseline → Save Baseline).",
    "   • Eksporter som .mpp og legg i 02 - Planlegging.",
    "",
    "5. Forventet resultat: en Gantt-plan med 116 oppgaver i 4 nivåer, korrekte FS-avhengigheter,",
    "   varigheter og kostnader, klar for status tracking i fase 3.",
]
for i, t in enumerate(steps, start=2):
    ws2.cell(row=i, column=1, value=t)
ws2.column_dimensions["A"].width = 120

wb.save(PATH)
print(f"Lagret: {PATH}")
print(f"Antall oppgaver: {len(W)}")
