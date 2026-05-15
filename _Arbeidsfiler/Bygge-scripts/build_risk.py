"""Bygger Risikoregister - Nye Hædda barneskole.xlsx med scoring, budsjett, restrisiko og eierskifte."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PATH = r"C:\Users\Gusta\OneDrive\Documents\Privat\Studier\Bachelore i Logistikk\LOG565 - Prosjektledelse\02 - Planlegging\Risikoregister - Nye Hædda barneskole.xlsx"

# (ID, Kategori, Beskrivelse, S, K, Eier, Tiltaksansv, Tiltak, Rest-S, Rest-K,
#  Budsjett tid (dager), Budsjett kost (mill kr), WBS-ref, Status)
risikoer = [
    ("R-001", "Grunnforhold",
     "Uforutsett kvikkleire eller fjell krever sprengning og forsterket fundamentering.",
     4, 5, "Prosjekteier", "Geotekniker",
     "Tidlig geoteknisk grunnundersøkelse (borplan før graving). Buffer i tidsplan og budsjett.",
     2, 3, 30, 6.0, "3.3", "Aktiv"),

    ("R-002", "Logistikk/Leverandør",
     "Forsinket levering av betongelementer eller spesialvinduer.",
     3, 4, "Innkjøpsleder", "Innkjøp",
     "Bestille tidlig, ha reserveleverandør, kontraktfeste leveringssanksjoner.",
     2, 2, 14, 1.5, "4.1.3", "Aktiv"),

    ("R-003", "HMS",
     "Alvorlig fall- eller klemulykke under arbeid i høyden.",
     2, 5, "HMS-leder", "Verneombud",
     "Daglig SJA, sertifisert sikringsutstyr, opplæring, vernerunder.",
     1, 4, 5, 0.5, "4.1", "Aktiv"),

    ("R-004", "Naboforhold",
     "Stopp i arbeid pga. støyklage fra naboer.",
     3, 3, "Prosjektleder", "PL/Kommunikasjon",
     "Fast info-runder, definerte arbeidstider, dempende tiltak ved tunge maskiner.",
     2, 2, 7, 0.4, "1.3.2", "Aktiv"),

    ("R-005", "Økonomi",
     "Kostnadsoverskridelse som følge av råvareprisøkning (stål, betong, glass).",
     4, 3, "Prosjekteier", "Innkjøpsleder",
     "Fastprisavtaler i NOK der mulig, indeksregulering med tak, reservere indeksbuffer.",
     3, 2, 0, 8.0, "4.1.1", "Aktiv"),

    ("R-006", "Myndighet",
     "Forsinket igangsettings- eller brukstillatelse fra kommune.",
     2, 4, "Prosjekteier", "Arkitekt/PL",
     "Tidlig forhåndskonferanse, løpende dialog med saksbehandler, komplett dokumentasjon.",
     1, 3, 21, 1.0, "2.2", "Aktiv"),

    ("R-007", "Teknisk",
     "Feil eller forsinket idriftsetting av SD-anlegg gir senere overtakelse.",
     3, 3, "Driftssjef", "Automasjonsentreprenør",
     "Tidlig integrasjonstest, kompetansekrav i kontrakt, prøvedrift før overtakelse.",
     2, 2, 14, 0.8, "5.4", "Aktiv"),

    ("R-008", "Ressurser/Personell",
     "Sykdom eller turnover i nøkkelroller (PL, RIB, RIE).",
     3, 3, "Prosjekteier", "Prosjektleder",
     "Backup-roller, kunnskapsoverføring, dokumentasjonsplikt, ekstern stedfortreder klar.",
     2, 2, 10, 0.6, "1.0", "Aktiv"),

    ("R-009", "Vær",
     "Ekstremvær (snø, frost, storm) stopper utendørs arbeid.",
     4, 3, "Entreprenør", "Anleggsleder",
     "Vinterdriftsplan, telt for kritiske arbeider, buffere i fremdrift.",
     3, 2, 14, 1.0, "4.1.2", "Aktiv"),

    ("R-010", "Endring/Omfang",
     "Mange sene endringsønsker fra skole/kommune (scope creep).",
     4, 3, "Prosjekteier", "Prosjektleder",
     "Etablere stoppdato for endringer, formell endringshåndtering med konsekvensanalyse.",
     3, 2, 7, 1.5, "1.1.3", "Aktiv"),

    ("R-011", "Brann/Sikkerhet",
     "Brann i byggefase med skade på pågående arbeid.",
     1, 5, "Brannrådgiver", "HMS-leder",
     "Tett-arbeider varmeplan, brannvakter, slukkeutstyr, ingen oppbevaring av brannlast.",
     1, 4, 21, 3.0, "4.1", "Aktiv"),

    ("R-012", "Akustikk",
     "Etterklangstid i klasserom over 0,5 sek (NS 8175 ikke møtt).",
     2, 3, "RIA", "Totalentreprenør",
     "Akustikkberegning i prosjektering, tidlig mock-up, måling før overtakelse.",
     1, 2, 7, 0.4, "4.2.1.3", "Aktiv"),

    ("R-013", "BREEAM",
     "BREEAM-poeng faller under Very Good (≥ 55 poeng).",
     2, 3, "Miljøkoordinator", "Miljøkoordinator",
     "Tidlig BREEAM-vurdering, milestone-oppfølging, rette opp underveis.",
     1, 2, 0, 0.8, "M-001", "Aktiv"),

    ("R-014", "IKT/Sikkerhet",
     "Mangelfull WiFi-dekning eller IKT-sikkerhet ved overtakelse.",
     2, 4, "Driftssjef", "IKT-leverandør",
     "Dekningsmåling og pen-test før overtakelse, krav om innregulering i kontrakt.",
     1, 3, 7, 0.5, "5.5", "Aktiv"),

    ("R-015", "Overtakelse",
     "Mer enn null kritiske mangler ved BP3 — utsatt overtakelse.",
     3, 4, "Prosjektleder", "Totalentreprenør",
     "Strukturert testprotokoll, pre-befaringer, mangelliste lukket før hovedbefaring.",
     2, 3, 14, 1.0, "8.2", "Aktiv"),
]

print(f"Antall risikoer: {len(risikoer)}")

# Beregn risikoscore og kategori
def kat_score(score):
    if score <= 4:
        return "Lav", "92D050"
    elif score <= 9:
        return "Middels", "FFD966"
    elif score <= 15:
        return "Høy", "F4B084"
    else:
        return "Kritisk", "FF7C80"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Risikoregister"

# Tittel
ws["A1"] = "RISIKOREGISTER — Nye Hædda barneskole"
ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.merge_cells("A1:O1")
ws.row_dimensions[1].height = 28

ws["A2"] = "Versjon 1.0 — 04.05.2026 — Sannsynlighet og konsekvens på skala 1–5. Score = S × K."
ws["A2"].font = Font(italic=True, size=10)
ws["A2"].alignment = Alignment(horizontal="center")
ws.merge_cells("A2:O2")

# Header
headers = [
    "ID", "Kategori", "Beskrivelse",
    "S (1-5)", "K (1-5)", "Score", "Risikonivå",
    "Risikoeier", "Tiltaksansvarlig", "Tiltak",
    "Rest-S", "Rest-K", "Rest-score",
    "Budsjett tid (dager)", "Budsjett kostnad (mill. kr)",
    "WBS-ref", "Status",
]
hr = 4
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=hr, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2E75B6")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

tot_tid = 0
tot_kost = 0.0
for i, r in enumerate(risikoer):
    rid, kat, beskr, S, K, eier, tans, tiltak, rS, rK, dager, kost, wbs, status = r
    score = S * K
    rscore = rS * rK
    nivaa, fcolor = kat_score(score)
    rest_nivaa, rest_fcolor = kat_score(rscore)
    row = hr + 1 + i
    vals = [rid, kat, beskr, S, K, score, nivaa, eier, tans, tiltak, rS, rK, rscore, dager, kost, wbs, status]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if c == 1:
            cell.font = Font(bold=True)
        if c in (6, 7):
            cell.fill = PatternFill("solid", fgColor=fcolor)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if c in (13,):
            cell.fill = PatternFill("solid", fgColor=rest_fcolor)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if c in (4, 5, 11, 12, 14, 15):
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tot_tid += dager
    tot_kost += kost

# Sum-rad
row = hr + 1 + len(risikoer)
ws.cell(row=row, column=13, value="SUM RISIKOBUDSJETT:").font = Font(bold=True)
ws.cell(row=row, column=13).alignment = Alignment(horizontal="right")
ws.cell(row=row, column=14, value=tot_tid).font = Font(bold=True)
ws.cell(row=row, column=15, value=round(tot_kost, 1)).font = Font(bold=True)
ws.cell(row=row, column=14).fill = PatternFill("solid", fgColor="FFE699")
ws.cell(row=row, column=15).fill = PatternFill("solid", fgColor="FFE699")
ws.cell(row=row, column=14).alignment = Alignment(horizontal="center")
ws.cell(row=row, column=15).alignment = Alignment(horizontal="center")

widths = [8, 18, 50, 8, 8, 8, 12, 20, 20, 55, 8, 8, 10, 18, 22, 12, 10]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A5"

# === Risikomatrise ===
ws2 = wb.create_sheet("Risikomatrise")
ws2["A1"] = "Risikomatrise — Nye Hædda barneskole"
ws2["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws2["A1"].fill = PatternFill("solid", fgColor="1F4E79")
ws2.merge_cells("A1:G1")

ws2["A3"] = "Fargekoding:"
ws2["A3"].font = Font(bold=True)
ws2.cell(row=3, column=2, value="Lav (1-4)").fill = PatternFill("solid", fgColor="92D050")
ws2.cell(row=3, column=3, value="Middels (5-9)").fill = PatternFill("solid", fgColor="FFD966")
ws2.cell(row=3, column=4, value="Høy (10-15)").fill = PatternFill("solid", fgColor="F4B084")
ws2.cell(row=3, column=5, value="Kritisk (16-25)").fill = PatternFill("solid", fgColor="FF7C80")

# Matrise (5x5) — sannsynlighet på Y, konsekvens på X
ws2["A6"] = "Sannsynlighet ↓ / Konsekvens →"
ws2["A6"].font = Font(bold=True)
for k in range(1, 6):
    ws2.cell(row=6, column=1+k, value=f"K={k}").font = Font(bold=True)
    ws2.cell(row=6, column=1+k).alignment = Alignment(horizontal="center")
for s in range(5, 0, -1):
    rr = 7 + (5 - s)
    ws2.cell(row=rr, column=1, value=f"S={s}").font = Font(bold=True)
    for k in range(1, 6):
        sc = s * k
        nv, fc = kat_score(sc)
        cell = ws2.cell(row=rr, column=1+k, value=sc)
        cell.fill = PatternFill("solid", fgColor=fc)
        cell.alignment = Alignment(horizontal="center")

# Plasser hver risiko i matrisen (vise hvilke ID-er ligger hvor)
ws2["A14"] = "Plassering før tiltak:"
ws2["A14"].font = Font(bold=True)
from collections import defaultdict
buckets = defaultdict(list)
for r in risikoer:
    buckets[(r[3], r[4])].append(r[0])
rr = 15
for (S, K), ids in sorted(buckets.items(), key=lambda x: -(x[0][0]*x[0][1])):
    ws2.cell(row=rr, column=1, value=f"S={S}, K={K}, Score={S*K}")
    ws2.cell(row=rr, column=2, value=", ".join(ids))
    rr += 1

for c in range(1, 8):
    ws2.column_dimensions[get_column_letter(c)].width = 14
ws2.column_dimensions["A"].width = 24

# === Sammendrag-ark ===
ws3 = wb.create_sheet("Sammendrag")
ws3["A1"] = "Risikoregister — sammendrag"
ws3["A1"].font = Font(bold=True, size=14)
ws3["A3"] = "Totalt antall risikoer:"
ws3["B3"] = len(risikoer)
ws3["A4"] = "Sum risikobudsjett — tid (dager):"
ws3["B4"] = tot_tid
ws3["A5"] = "Sum risikobudsjett — kostnad (mill. kr):"
ws3["B5"] = round(tot_kost, 1)

ws3["A7"] = "Fordeling pr nivå (før tiltak):"
ws3["A7"].font = Font(bold=True)
from collections import Counter
levels = Counter()
for r in risikoer:
    nv, _ = kat_score(r[3]*r[4])
    levels[nv] += 1
rr = 8
for k in ["Kritisk", "Høy", "Middels", "Lav"]:
    ws3.cell(row=rr, column=1, value=k)
    ws3.cell(row=rr, column=2, value=levels.get(k, 0))
    rr += 1

ws3["A14"] = "Fordeling pr nivå (etter tiltak):"
ws3["A14"].font = Font(bold=True)
levels_r = Counter()
for r in risikoer:
    nv, _ = kat_score(r[8]*r[9])
    levels_r[nv] += 1
rr = 15
for k in ["Kritisk", "Høy", "Middels", "Lav"]:
    ws3.cell(row=rr, column=1, value=k)
    ws3.cell(row=rr, column=2, value=levels_r.get(k, 0))
    rr += 1

ws3.column_dimensions["A"].width = 42
ws3.column_dimensions["B"].width = 14

# === Til Bård-ark ===
ws4 = wb.create_sheet("Til Bård")
ws4["A1"] = "Til Bård — om risikoregisteret"
ws4["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws4["A1"].fill = PatternFill("solid", fgColor="1F4E79")
ws4.merge_cells("A1:F1")
notes = [
    "",
    "Hei Bård,",
    "",
    f"Vi har {len(risikoer)} risikoer registrert med scoring 1-5 på sannsynlighet og konsekvens.",
    "Hver risiko er knyttet til relevant WBS-element via 'WBS-ref'.",
    "",
    f"Foreslått risikobudsjett (totalt over alle risikoer):",
    f"  • Tid: {tot_tid} dager",
    f"  • Kostnad: {round(tot_kost,1)} mill. kr",
    "",
    "Vi setter pris på dine kommentarer/justeringer hvis noe er urealistisk satt.",
    "",
    "Mvh",
    "Studentgruppe LOG565",
]
for i, t in enumerate(notes, start=2):
    ws4.cell(row=i, column=1, value=t)
ws4.column_dimensions["A"].width = 100

wb.save(PATH)
print(f"Lagret: {PATH}")
