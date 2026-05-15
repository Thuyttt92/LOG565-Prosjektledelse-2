"""Bygger WBS - Nye Hædda barneskole.xlsx med 65+ leveranser i 4 nivåer.

Tomme kolonner Varighet (mnd), Kostnadsestimat (mill. kr) og Estimat fra Bård
fylles av Bård når han returnerer filen.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PATH = r"C:\Users\Gusta\OneDrive\Documents\Privat\Studier\Bachelore i Logistikk\LOG565 - Prosjektledelse\02 - Planlegging\WBS - Nye Hædda barneskole.xlsx"

# Format: (WBS-ID, Nivå, Aktivitet/leveranse, Beskrivelse, Ansvarlig, Avhengigheter, Krav-ID)
W = []

def add(wid, niv, navn, beskr, ansv, avh="", krav=""):
    W.append((wid, niv, navn, beskr, ansv, avh, krav))

# === 1 PROSJEKTLEDELSE ===
add("1", 1, "Prosjektledelse og administrasjon", "Overordnet styring og administrasjon gjennom hele prosjektløpet.", "Prosjektleder")
add("1.1", 2, "Prosjektstyring", "Fremdriftsplanlegging, økonomistyring og rapportering.", "Prosjektleder")
add("1.1.1", 3, "Etablere baseline (Gantt + budsjett)", "Låse plan og kostnadsbase etter Bårds estimater.", "Prosjektleder")
add("1.1.2", 3, "Månedlig styringsrapportering", "Statusrapport, KPI, S-kurve, EVM hver måned.", "Prosjektleder")
add("1.1.3", 3, "Endringsstyring", "Behandle endringsforespørsler, konsekvensanalyse, godkjenning.", "Prosjektleder")
add("1.2", 2, "Kontraktsoppfølging", "Oppfølging av totalentreprenør (NS 8407) og leverandører.", "Prosjektleder")
add("1.2.1", 3, "Kontraktsforhandling", "Etablere kontrakter med entreprenører.", "Prosjektleder/Innkjøp")
add("1.2.2", 3, "Avviks- og endringslogg", "Løpende logg over avvik mot kontrakt.", "Prosjektleder")
add("1.3", 2, "Interessenthåndtering", "Koordinering med skoleledelse, ansatte, foreldre, kommune.", "Prosjektleder", krav="F-013")
add("1.3.1", 3, "Brukermøter", "Faste møter med skole og kommune.", "Prosjektleder")
add("1.3.2", 3, "Kommunikasjonsplan og nyhetsbrev", "Informasjon til naboer og presse.", "Prosjektleder", krav="R-004")
add("1.4", 2, "Risiko- og kvalitetsstyring", "Risikoregister, HMS-kontroller og kvalitetssikring.", "HMS-/KS-ansvarlig")
add("1.4.1", 3, "Vedlikehold av risikoregister", "Månedlig oppdatering av risikoregister.", "HMS-/KS-ansvarlig")
add("1.4.2", 3, "HMS-revisjoner og SJA", "Daglige sikker-jobb-analyser i utførelsen.", "HMS-leder", krav="R-003")
add("1.4.3", 3, "Kvalitetsplan og kontroller", "Etablere KS-plan, gjennomføre revisjoner.", "KS-ansvarlig", krav="K-001")

# === 2 PLANLEGGING/PROSJEKTERING ===
add("2", 1, "Planlegging og prosjektering", "Detaljprosjektering, godkjenninger og forberedelse til konkurranse.", "Prosjekteringsleder")
add("2.1", 2, "Detaljprosjektering", "Arkitekt + tekniske fag.", "Arkitekt")
add("2.1.1", 3, "Arkitekttegninger og romprogram", "Tegninger med detaljert romprogram.", "Arkitekt", krav="F-001;F-004")
add("2.1.2", 3, "Konstruksjonstegninger (RIB)", "Bærekonstruksjoner, statisk beregning.", "RIB", krav="T-001")
add("2.1.3", 3, "VVS-tegninger (RIV)", "Varme, ventilasjon, sanitær.", "RIV", krav="T-002;T-003")
add("2.1.4", 3, "Elektro-tegninger (RIE)", "Kraft, belysning, sterkstrøm.", "RIE", krav="T-008")
add("2.1.5", 3, "IKT- og sikkerhetsprosjektering", "Nettverk, AV, adgangskontroll.", "IKT-rådgiver", krav="T-005;S-003")
add("2.1.6", 3, "Akustisk prosjektering (RIA)", "Romakustikk og lydskillevegg.", "RIA", krav="A-001;A-002;A-003")
add("2.2", 2, "Offentlige godkjenninger", "Søknader til kommunen.", "Arkitekt/PL")
add("2.2.1", 3, "Rammetillatelse", "Søknad og oppfølging av rammetillatelse.", "Arkitekt", krav="R-006")
add("2.2.2", 3, "Igangsettingstillatelse (IG)", "Etappe-IG for graving og bygg.", "Arkitekt", avh="2.2.1", krav="R-006")
add("2.2.3", 3, "Nabovarsel og innsigelseshåndtering", "Formell varsling og dialog.", "Arkitekt", krav="R-004")
add("2.3", 2, "Konkurransegrunnlag", "Kontrakter og krav til entreprenører.", "Innkjøpsleder")
add("2.3.1", 3, "Kontraktstrategi og kvalifikasjonskrav", "Velge kontraktsform (NS 8407), lage tilbudsgrunnlag.", "Innkjøpsleder")
add("2.3.2", 3, "Tilbudsfase og kontrahering", "Innhente og evaluere tilbud, signere kontrakt.", "Innkjøpsleder")

# === 3 FORBEREDELSE / RIVING ===
add("3", 1, "Forberedelse og riving", "Klargjøring av tomt: miljøsanering, riving, grunnarbeid.", "Entreprenør", avh="2")
add("3.1", 2, "Miljøsanering", "Kartlegging og fjerning av miljøskadelige stoffer.", "Miljørådgiver")
add("3.1.1", 3, "Miljøkartlegging", "Befaring og prøvetaking av eksisterende bygg.", "Miljørådgiver")
add("3.1.2", 3, "Sanering og avhending", "Fjerning og deponering iht. forskrift.", "Miljøentreprenør", avh="3.1.1", krav="M-006")
add("3.2", 2, "Riving av eksisterende skole", "Fysisk fjerning av eksisterende bygningsmasse.", "Entreprenør", avh="3.1")
add("3.3", 2, "Grunnarbeid og fundamentering", "Geotekniske arbeider og fundamentering.", "Grunnentreprenør", avh="3.2")
add("3.3.1", 3, "Geoteknisk grunnundersøkelse", "Borprøver og rapport.", "Geotekniker", krav="R-001")
add("3.3.1.1", 4, "Borplan og prøvetaking", "Plassering og uttak av borprøver.", "Geotekniker")
add("3.3.1.2", 4, "Geoteknisk rapport", "Anbefaling for fundamentering.", "Geotekniker")
add("3.3.2", 3, "Masseutskifting og sprenging", "Fjerne dårlige masser og fjell.", "Grunnentreprenør", avh="3.3.1")
add("3.3.3", 3, "Etablere fundamenter og bunnplate", "Støping av plate på mark.", "Grunnentreprenør", avh="3.3.2")

# === 4 SKOLEBYGG (BYGNINGSMESSIG) ===
add("4", 1, "Skolebygg — bygningsmessige arbeider", "Råbygg og innvendig komplettering over 3 etasjer.", "Totalentreprenør", avh="3")
add("4.1", 2, "Råbygg", "Bærekonstruksjoner, tak og fasade.", "Totalentreprenør")
add("4.1.1", 3, "Bærekonstruksjon", "Søyler, dragere, dekker.", "Totalentreprenør", krav="T-001")
add("4.1.1.1", 4, "Søyler og dragere (stål/limtre)", "Reisning av primærbæring.", "Totalentreprenør")
add("4.1.1.2", 4, "Etasjeskillere (hulldekker)", "Montering av hulldekker plan 1–3.", "Totalentreprenør")
add("4.1.1.3", 4, "Stabiliserende vegger og avstivninger", "Bjelker og avstivende vegger.", "Totalentreprenør")
add("4.1.2", 3, "Tak og takmembran inkl. sedumtak", "Tett tak med grønn taklag.", "Totalentreprenør", avh="4.1.1", krav="M-002")
add("4.1.2.1", 4, "Taktekking og membran", "Tett tak.", "Totalentreprenør")
add("4.1.2.2", 4, "Sedumlag og taksluk", "Grønn taklag og overvannshåndtering.", "Totalentreprenør", krav="M-002")
add("4.1.3", 3, "Fasade og vinduer", "Yttervegg, isolasjon, vinduer og solavskjerming.", "Totalentreprenør", avh="4.1.1")
add("4.1.3.1", 4, "Yttervegger og isolasjon", "Bindingsverk, mineralull, vindsperre.", "Totalentreprenør")
add("4.1.3.2", 4, "Vinduer og glassfelt", "U-verdi ≤ 0,8.", "Totalentreprenør")
add("4.1.3.3", 4, "Fasadekledning og solavskjerming", "Trekledning + utvendig solskjerming.", "Totalentreprenør")
add("4.2", 2, "Innvendig 1. etasje (småtrinn 1–4)", "Klasserom 1–4, SFO, helse, kantine, gymsalbirom.", "Totalentreprenør", avh="4.1")
add("4.2.1", 3, "Klasseromsfløy 1–4", "Innredning og overflater for småtrinn.", "Totalentreprenør", krav="F-002;F-003;A-001")
add("4.2.1.1", 4, "Lettvegger og dører", "Lettvegger med stålstendere, branndører.", "Totalentreprenør")
add("4.2.1.2", 4, "Gulv (slitesterk belegg)", "Vinyl/linoleum klasse 33.", "Totalentreprenør", krav="D-001")
add("4.2.1.3", 4, "Vegger og himling med akustikkplater", "Maling og akustikkbehandling.", "Totalentreprenør", krav="A-001")
add("4.2.2", 3, "SFO-areal med separat inngang", "Eget område med direkte uteforbindelse.", "Totalentreprenør", krav="F-013")
add("4.2.3", 3, "Helseavdeling og kantine", "Skolelegerom, kantineareal og servering.", "Totalentreprenør", krav="F-008;F-011")
add("4.3", 2, "Innvendig 2. etasje (mellomtrinn 5–7)", "Klasserom 5–7, bibliotek, auditorium, administrasjon.", "Totalentreprenør", avh="4.1")
add("4.3.1", 3, "Klasseromsfløy 5–7", "Innredning mellomtrinn.", "Totalentreprenør", krav="F-002;F-003;A-001")
add("4.3.2", 3, "Bibliotek og auditorium", "Møblering, AV-utstyr, akustikk.", "Totalentreprenør", krav="F-005;A-003")
add("4.3.3", 3, "Administrasjon og resepsjon", "Resepsjon, kontorer, møterom.", "Totalentreprenør", krav="F-009;F-015")
add("4.4", 2, "Innvendig 3. etasje (ungdomstrinn 8–10)", "Klasserom 8–10, spesialrom, drift.", "Totalentreprenør", avh="4.1")
add("4.4.1", 3, "Klasseromsfløy 8–10", "Innredning ungdomstrinn.", "Totalentreprenør", krav="F-002;F-003;A-001")
add("4.4.2", 3, "Spesialrom (naturfag, musikk, K&H, mat&helse)", "Spesialinnredning og utstyrspunkt.", "Totalentreprenør", krav="F-006")
add("4.4.3", 3, "Driftssone (vaktmester, lager, verksted)", "Adskilt driftssone.", "Totalentreprenør", krav="F-010")
add("4.5", 2, "Gymsal", "Sentralt plassert gymsal med birom.", "Totalentreprenør", avh="4.1")
add("4.5.1", 3, "Hovedhall og sportsgulv", "Slitesterkt sportsgulv, vegglinjer.", "Totalentreprenør", krav="F-007;A-002")
add("4.5.2", 3, "Garderober og dusjanlegg", "Tilstrekkelig kapasitet for samtidig bruk.", "Totalentreprenør", krav="F-012")

# === 5 TEKNISKE ANLEGG ===
add("5", 1, "Skolebygg — tekniske anlegg", "VVS, elektro, heis, automasjon, IKT/sikkerhet.", "Tekniske entreprenører", avh="4.1")
add("5.1", 2, "VVS", "Varme, ventilasjon, sanitær, sprinkler.", "VVS-entreprenør")
add("5.1.1", 3, "Vannbåren varme + fjernvarme", "Sentral varme med tilkobling.", "VVS-entreprenør", krav="T-004")
add("5.1.2", 3, "Balansert ventilasjon m/varmegjenvinning", "Aggregat og kanalføring.", "VVS-entreprenør", krav="T-002;T-003")
add("5.1.2.1", 4, "Ventilasjonsaggregat på tak", "Montering og elektrisk tilkobling.", "VVS-entreprenør")
add("5.1.2.2", 4, "Hovedkanaler og fordeling", "Sjakter og hovedkanaler i himling.", "VVS-entreprenør")
add("5.1.2.3", 4, "Tilluft- og avtrekksventiler", "Endepunkter i hvert rom.", "VVS-entreprenør")
add("5.1.3", 3, "Sanitær- og avløpsanlegg", "Toaletter, dusj, kjøkken, avløp.", "VVS-entreprenør", krav="F-012")
add("5.1.4", 3, "Sprinkleranlegg", "Sentral, ledningsnett og hoder.", "VVS/Brann", krav="S-001")
add("5.2", 2, "Elektro", "Kraft, belysning, lynvern, UPS.", "Elektroentreprenør")
add("5.2.1", 3, "Hovedfordeling og kraftkurser", "Inntak, hovedtavle, fordelinger.", "Elektroentreprenør", krav="T-009")
add("5.2.2", 3, "Belysning (LED + styring)", "LED, dagslys-/tilstedeværelsesstyring.", "Elektroentreprenør", krav="T-008")
add("5.2.3", 3, "UPS og nødlys", "Reservekraft til IKT og nødlys.", "Elektroentreprenør", krav="T-009")
add("5.2.4", 3, "Lynvern og jordingsanlegg", "Ekstern og intern lynvern.", "Elektroentreprenør")
add("5.3", 2, "Heis og vertikal transport", "Personheis mellom alle etasjer.", "Heisleverandør", krav="T-007;UU-002")
add("5.4", 2, "Automasjon — SD-anlegg", "Sentral driftskontroll for HVAC, lys, adgang.", "Automasjonsentreprenør", krav="T-006;D-002")
add("5.4.1", 3, "Programmering og tagliste", "Definere taglister og styringssløyfer.", "Automasjonsentreprenør")
add("5.4.2", 3, "Idriftsetting og innregulering", "Test og innstilling av regulering.", "Automasjonsentreprenør", avh="5.4.1")
add("5.5", 2, "IKT og sikkerhet", "Nettverk, adgangskontroll, kamera, alarm, AV.", "IKT-leverandør")
add("5.5.1", 3, "Strukturert kabling og WiFi", "Spredenett og trådløst nett.", "IKT-leverandør", krav="T-005")
add("5.5.1.1", 4, "Cat 6A-kabling og patcheskap", "Spredenett til alle uttak.", "IKT-leverandør")
add("5.5.1.2", 4, "WiFi-aksesspunkt", "Dekning ≥ 100 Mbps i alle rom.", "IKT-leverandør", krav="T-005")
add("5.5.1.3", 4, "Innregulering og dekningsmåling", "Sluttest og rapport.", "IKT-leverandør")
add("5.5.2", 3, "Adgangskontroll og lockdown", "Kortlesere, sentral lockdown.", "IKT/Elektro", krav="S-003;S-006")
add("5.5.3", 3, "Kamera- og innbruddsalarm", "Kamerasystem og alarmsoner.", "IKT", krav="S-004;S-005")
add("5.5.4", 3, "AV-utstyr i klasserom og auditorium", "Skjermer, lyd, presentasjonssystem.", "IKT", krav="T-010")
add("5.5.5", 3, "Serverrom og teleslynge", "Kjølt serverrom + teleslynge i auditorium.", "IKT", krav="T-011;UU-004")

# === 6 UTOMHUS ===
add("6", 1, "Utomhus og uteområder", "Lekearealer, sport, infrastruktur, grønt.", "Landskapsentreprenør", avh="3")
add("6.1", 2, "Lekearealer", "Aldersinndelt lek med fallunderlag.", "Landskapsentreprenør", krav="U-001;U-006")
add("6.2", 2, "Sport og fritid", "Ballbinge og sosiale soner med utebelysning.", "Landskapsentreprenør", krav="U-002;U-004")
add("6.3", 2, "Infrastruktur og parkering", "Veier, P-plasser, sykkel, ladeplasser, hentesoner.", "Anleggsentreprenør", krav="M-003;U-003;U-007")
add("6.4", 2, "Grøntanlegg og støyskjerming", "Beplantning og skjerming mot vei.", "Landskapsentreprenør", krav="U-005")

# === 7 INVENTAR / FF&E ===
add("7", 1, "Inventar og utstyr (FF&E)", "Møbler, spesialutstyr, AV-løsninger.", "Innkjøpsleder", avh="4")
add("7.1", 2, "Løst inventar", "Møbler i klasserom, kontorer, fellesarealer.", "Innkjøpsleder")
add("7.2", 2, "Spesialutstyr", "Naturfag, musikk, K&H, mat & helse.", "Innkjøpsleder", krav="F-006")
add("7.3", 2, "AV-løsninger", "Skjermer, lydanlegg, presentasjon.", "IKT-leverandør", krav="T-010")

# === 8 OVERTAKELSE / AVSLUTNING ===
add("8", 1, "Overtakelse og avslutning", "Sluttkontroll, overlevering, opplæring, evaluering.", "Prosjektleder", avh="5;7")
add("8.1", 2, "Testing og prøvedrift", "Test av alle tekniske anlegg.", "Prosjektleder")
add("8.1.1", 3, "Funksjonstester (HVAC, elektro, IKT)", "Strukturert testprotokoll.", "Tekniske entreprenører")
add("8.1.2", 3, "Brann- og rømningstest", "Evakueringsøvelse + brannprøve.", "Brannrådgiver", krav="S-002")
add("8.2", 2, "Ferdigbefaring (BP3)", "Sluttkontroll og lukke mangelliste.", "Prosjektleder", avh="8.1", krav="K-001")
add("8.3", 2, "FDV-dokumentasjon", "Komplett FDV-pakke digitalt.", "FDV-ansvarlig", krav="K-002")
add("8.4", 2, "Brukeropplæring", "Opplæring av drifts- og pedagogisk personell.", "Driftssjef", avh="8.2", krav="D-004")
add("8.5", 2, "Prosjektevaluering og sluttrapport", "Erfaringshøsting og endelig sluttrapport.", "Prosjektleder", avh="8.2;8.3;8.4")

print(f"Antall WBS-linjer: {len(W)}")
nivaaer = set(w[1] for w in W)
print(f"Nivåer: {sorted(nivaaer)}")
n_lev = sum(1 for w in W if w[1] >= 2)
print(f"Antall leveranser (nivå >= 2): {n_lev}")
assert len(W) >= 65, "For få WBS-linjer"
assert max(nivaaer) >= 3, "WBS når ikke nok dybde"

# === Bygg Excel-fil ===
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "WBS"

ws["A1"] = "WBS — Nye Hædda barneskole"
ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.merge_cells("A1:K1")
ws.row_dimensions[1].height = 28

ws["A2"] = "Versjon 1.0 — 04.05.2026 — Sendt til Bård for tids- og kostnadsestimater"
ws["A2"].font = Font(italic=True, size=10)
ws["A2"].alignment = Alignment(horizontal="center")
ws.merge_cells("A2:K2")

# Bård fyller inn de tre siste kolonnene
headers = ["WBS-ID", "Nivå", "Aktivitet / leveranse", "Beskrivelse", "Ansvarlig", "Avhengigheter", "Krav-ID (sporbarhet)",
           "Varighet (mnd) — Bård", "Start — Bård", "Slutt — Bård", "Kostnadsestimat (mill. kr) — Bård"]
hr = 4
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=hr, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2E75B6") if c <= 7 else PatternFill("solid", fgColor="C65911")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

niv_colors = {1: "1F4E79", 2: "9DC3E6", 3: "DEEBF7", 4: "F2F2F2"}
niv_text = {1: "FFFFFF", 2: "000000", 3: "000000", 4: "000000"}

for i, (wid, niv, navn, beskr, ansv, avh, krav) in enumerate(W):
    r = hr + 1 + i
    indent = "    " * (niv - 1)
    vals = [wid, niv, indent + navn, beskr, ansv, avh, krav, "", "", "", ""]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if niv == 1:
            cell.font = Font(bold=True, color=niv_text[niv])
            cell.fill = PatternFill("solid", fgColor=niv_colors[niv])
        elif niv == 2:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor=niv_colors[niv])
        else:
            cell.fill = PatternFill("solid", fgColor=niv_colors[niv])
        if c >= 8:
            # Bård fyller inn — gul tone
            cell.fill = PatternFill("solid", fgColor="FFF2CC")

widths = [10, 7, 50, 55, 22, 18, 22, 16, 14, 14, 22]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A5"

# Statistikk-ark
ws2 = wb.create_sheet("Statistikk")
ws2["A1"] = "Statistikk for WBS"
ws2["A1"].font = Font(bold=True, size=14)
ws2["A3"] = "Nivå"
ws2["B3"] = "Antall linjer"
ws2["A3"].font = Font(bold=True); ws2["B3"].font = Font(bold=True)
from collections import Counter
cnt = Counter(w[1] for w in W)
r = 4
for k in sorted(cnt.keys()):
    ws2.cell(row=r, column=1, value=f"Nivå {k}")
    ws2.cell(row=r, column=2, value=cnt[k])
    r += 1
ws2.cell(row=r, column=1, value="TOTALT").font = Font(bold=True)
ws2.cell(row=r, column=2, value=len(W)).font = Font(bold=True)
ws2["A1"].alignment = Alignment(horizontal="left")
ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 18
ws2["A8"] = "A-nivå krever ≥ 60 leveranser og 4 nivåer."
ws2["A8"].font = Font(italic=True)

# Til Bård-ark
ws3 = wb.create_sheet("Til Bård")
ws3["A1"] = "Til Bård — instruksjoner for estimering"
ws3["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws3["A1"].fill = PatternFill("solid", fgColor="1F4E79")
ws3.merge_cells("A1:F1")
notes = [
    "",
    "Hei Bård,",
    "",
    "Vi har bygget WBS-en i 4 nivåer med totalt %d linjer (mål for A: ≥ 60 leveranser, 4 nivåer)." % len(W),
    "",
    "Vennligst fyll inn de gule kolonnene H–K i WBS-arket:",
    "  • H — Varighet i måneder (eller del-måneder, f.eks. 0,5)",
    "  • I — Estimert startmåned (f.eks. \"Februar 2025\")",
    "  • J — Estimert sluttmåned",
    "  • K — Kostnadsestimat i millioner kroner",
    "",
    "Hvert WBS-element er knyttet opp mot kravspesifikasjonen via kolonne G (Krav-ID).",
    "Avhengigheter ligger i kolonne F.",
    "",
    "Når estimatene er klare, sender vi en personlig melding på Teams.",
    "",
    "Ta gjerne kontakt på Teams hvis noe er uklart.",
    "",
    "Mvh",
    "Studentgruppe LOG565",
]
for i, t in enumerate(notes, start=2):
    ws3.cell(row=i, column=1, value=t)
ws3.column_dimensions["A"].width = 110

wb.save(PATH)
print(f"Lagret: {PATH}")
