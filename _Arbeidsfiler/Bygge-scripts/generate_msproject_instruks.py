"""
Genererer en omfattende MS Project tracking-instruks som xlsx:
  - Sheet 1: Steg-for-steg instruks (Baseline 0 → crashing → Baseline 1 → tracking)
  - Sheet 2: Baseline 1 oppgaver (32 leaves + 8 summary tasks med varighet, start/slutt, BAC)
  - Sheet 3: Tracking-data per måned (% complete + faktisk påløp per pakke per statusdato)
  - Sheet 4: Crashing-endringer (4.1 Råbygg pre vs post)

Brukeren bruker dette som referanse mens hen jobber i MS Project (manuelt).
"""
from __future__ import annotations
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from log565_master_data import (
    BAC, MÅNEDER, PAKKER, PAKKE_BY_WBS,
    AC_KUM, FREMDRIFT_PER_MND, HENDELSER,
    hent_pct_fullført,
)

sys.path.insert(0, str(Path(__file__).resolve().parent))
from paths import GJENNOMFORING
OUT = GJENNOMFORING / "MS Project tracking-instruks.xlsx"

# Styling
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SUBHDR_FILL = PatternFill("solid", fgColor="BDD7EE")
SUBHDR_FONT = Font(name="Calibri", size=10, bold=True, color="1F4E78")
ACCENT_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def style_header_row(ws, row, end_col):
    for c in range(1, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def build_instruks(wb):
    ws = wb.active
    ws.title = "Instruks"
    ws["A1"] = "MS Project Tracking-instruks — Nye Hædda Barneskole"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
    ws.merge_cells("A1:B1")

    ws["A2"] = ("Følg disse stegene i MS Project for å reflektere hele endringshistorikken "
                "(Baseline 0 → crashing → Baseline 1 → faktisk gjennomføring). Eksakte verdier "
                "ligger i de andre arkene.")
    ws["A2"].font = Font(italic=True, color="666666")
    ws["A2"].alignment = WRAP
    ws.merge_cells("A2:B2")
    ws.row_dimensions[2].height = 35

    steg = [
        ("FORBEREDELSE", ""),
        ("0", "Åpne en kopi av 'Hædda barneskole GANTT.mpp' (fase 2). "
              "Lagre som 'Hædda barneskole GANTT - tracking.mpp' for å bevare originalen."),
        ("", ""),
        ("DEL A — BASELINE 0 (opprinnelig plan før crashing)", ""),
        ("A1", "Bekreft at WBS i Project speiler de 32 arbeidspakkene i Sheet 'Baseline 1 oppgaver' "
               "(samme WBS-koder 1.1–8.5). Hvis 4.1 Råbygg i Project har 7 måneders varighet og "
               "190 MNOK kostnad — det er Baseline 0-tilstanden (Bårds WBS-simulert)."),
        ("A2", "Sett Baseline 0: Prosjekt → Set Baseline → Baseline 0 → Entire project → OK. "
               "Dette fryser den opprinnelige planen før crashing."),
        ("A3", "Lagre prosjektet."),
        ("", ""),
        ("DEL B — APPLISER CRASHING (NHB-2026-15-vedtak)", ""),
        ("B1", "Åpne arbeidspakke 4.1 Råbygg. Endre varighet fra 7 måneder (140 dager) til "
               "5 måneder (100 dager)."),
        ("B2", "Endre kostnad på 4.1 Råbygg fra 190 MNOK til 240 MNOK "
               "(+50 MNOK = crashing-merkostnad iht. godkjenning-av-budsjettendring.pdf)."),
        ("B3", "Sluttdato på 4.1 Råbygg skal nå vise desember 2025 (måned 11). "
               "Verifiser at etterfølgende pakker (4.2/4.3/4.4/4.5, 5.x) starter januar 2026 (mnd 12) "
               "og prosjektets samlede sluttdato faller på 15. mai 2026."),
        ("B4", "Hvis totalkostnad for prosjektet nå viser 800 MNOK — perfekt. "
               "Hvis ikke, sjekk at de andre 31 pakkene har sin opprinnelige kostnad fra "
               "Bårds WBS-simulert (sum = 560 MNOK; 4.1 utgjør 240 MNOK; totalt 800 MNOK)."),
        ("", ""),
        ("DEL C — BASELINE 1 (godkjent etter crashing)", ""),
        ("C1", "Sett Baseline 1: Prosjekt → Set Baseline → Baseline 1 → Entire project → OK."),
        ("C2", "Lagre prosjektet. Behold begge baselinene — sluttrapporten viser begge."),
        ("C3", "I View → Tracking Gantt: legg til Baseline 0-stolper og Baseline 1-stolper "
               "(høyreklikk på timeline → Bar Styles → ny rad for Baseline 0). "
               "Da kan du eksportere et bilde som viser hele endringshistorikken."),
        ("", ""),
        ("DEL D — FAKTISK FREMDRIFT OG KOSTNAD (16 statusdatoer)", ""),
        ("D1", "Åpne View → Tracking Gantt."),
        ("D2", "For hver av de 16 statusdatoene (se ark 'Tracking per måned'):"),
        ("D2a", "  a) Project → Status Date → sett til siste dag i måneden (f.eks. 28.02.2025 for mnd 1)."),
        ("D2b", "  b) For hver pakke som er aktiv i den måneden: dobbeltklikk → fane 'Tracking' eller "
                "'General' → fyll inn % Complete (fra arket) og Actual Cost (kumulativ påløp = sum av "
                "perioden + tidligere). Eller fyll inn 'Actual Cost' for hver periode direkte i tabellen."),
        ("D2c", "  c) Lagre prosjektet. Gjenta for neste måned."),
        ("D3", "Etter mnd 16: total Actual Cost = 800 MNOK, alle pakker 100 % fullført, "
               "sluttdato 15.05.2026. CPI = SPI = 1.000."),
        ("", ""),
        ("DEL E — RAPPORTERING (EVM + GANTT-EKSPORT)", ""),
        ("E1", "Project → Reports → Costs → Earned Value Report. Verifiser at PV/EV/AC "
               "stemmer med S-kurven i månedsrapportene."),
        ("E2", "For hver månedsrapport: eksporter Tracking Gantt som bilde "
               "(View → Tracking Gantt → File → Print → Save as PDF, eller Snipping Tool på skjermbildet) "
               "med statuslinje på riktig dato. Lim inn under 'Figur 1' i månedsrapporten."),
        ("E3", "For sluttrapporten: eksporter en sammenstilt Gantt som viser "
               "Baseline 0 vs Baseline 1 vs faktisk."),
    ]
    r = 4
    for steg_nr, tekst in steg:
        if steg_nr and not tekst:
            # Header-rad
            ws.cell(row=r, column=1, value=steg_nr)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            cell = ws.cell(row=r, column=1)
            cell.fill = SUBHDR_FILL
            cell.font = SUBHDR_FONT
        elif steg_nr or tekst:
            ws.cell(row=r, column=1, value=steg_nr).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=1).font = Font(bold=True)
            ws.cell(row=r, column=2, value=tekst).alignment = WRAP
            ws.row_dimensions[r].height = max(20, 15 * (len(tekst) // 90 + 1))
        r += 1

    set_col_widths(ws, [10, 110])


def build_baseline1_tasks(wb):
    """Eksakte oppgaver for Baseline 1 (etter crashing)."""
    ws = wb.create_sheet("Baseline 1 oppgaver")
    ws["A1"] = "Baseline 1 — alle 32 arbeidspakker (etter crashing av 4.1 Råbygg)"
    ws["A1"].font = Font(size=14, bold=True, color="1F4E78")
    ws.merge_cells("A1:H1")

    ws["A2"] = ("Disse verdiene skal stemme i MS Project etter Del A–C i instruksen. "
                "Sum BAC = 800 MNOK. Sluttdato = 15.05.2026.")
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:H2")

    headers = ["WBS", "Aktivitet", "Varighet (mnd)", "Start (mnd)", "Slutt (mnd)",
               "BAC (MNOK)", "BAC (kr)", "Fagansvar"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, len(headers))

    for i, p in enumerate(PAKKER, start=5):
        start_navn, _ = MÅNEDER[p.start_mnd]
        slutt_navn, _ = MÅNEDER[p.slutt_mnd]
        ws.cell(row=i, column=1, value=p.wbs)
        ws.cell(row=i, column=2, value=p.navn)
        ws.cell(row=i, column=3, value=p.varighet).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=4, value=f"Mnd {p.start_mnd} — {start_navn}")
        ws.cell(row=i, column=5, value=f"Mnd {p.slutt_mnd} — {slutt_navn}")
        ws.cell(row=i, column=6, value=p.bac).number_format = '0.0'
        ws.cell(row=i, column=7, value=p.bac * 1_000_000).number_format = '#,##0'
        ws.cell(row=i, column=8, value=p.fagansvar)
        for c in range(1, 9):
            ws.cell(row=i, column=c).border = BORDER
        if p.wbs == "4.1":
            for c in range(1, 9):
                ws.cell(row=i, column=c).fill = ACCENT_FILL

    # SUM rad
    sum_r = 5 + len(PAKKER)
    ws.cell(row=sum_r, column=2, value="TOTAL").font = Font(bold=True)
    ws.cell(row=sum_r, column=6, value=f"=SUM(F5:F{sum_r-1})").number_format = '0.0'
    ws.cell(row=sum_r, column=6).font = Font(bold=True)
    ws.cell(row=sum_r, column=7, value=f"=SUM(G5:G{sum_r-1})").number_format = '#,##0'
    ws.cell(row=sum_r, column=7).font = Font(bold=True)

    set_col_widths(ws, [8, 40, 12, 22, 22, 12, 16, 28])


def build_crashing_endring(wb):
    ws = wb.create_sheet("Crashing-endring")
    ws["A1"] = "Crashing-endring på 4.1 Råbygg (NHB-2026-15)"
    ws["A1"].font = Font(size=14, bold=True, color="1F4E78")
    ws.merge_cells("A1:D1")

    ws["A3"] = "Aktivitet 4.1 Råbygg endres iht. kommunestyrets vedtak datert 07.05.2026."
    ws["A3"].font = Font(italic=True)
    ws.merge_cells("A3:D3")

    headers = ["Parameter", "Baseline 0 (Bårds simulering)", "Baseline 1 (etter crashing)", "Endring"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=5, column=i, value=h)
    style_header_row(ws, 5, len(headers))

    rows = [
        ("Varighet", "7 måneder", "5 måneder", "–2 mnd"),
        ("Start", "Aug 2025 (mnd 7)", "Aug 2025 (mnd 7)", "Uendret"),
        ("Slutt", "Feb 2026 (mnd 13)", "Des 2025 (mnd 11)", "–2 mnd tidligere"),
        ("Kostnad", "190 MNOK", "240 MNOK", "+50 MNOK"),
        ("Prosjektets totalkostnad", "750 MNOK", "800 MNOK", "+50 MNOK (utvidet ramme)"),
        ("Prosjektets sluttdato", "Juli 2026 (forsinket)", "15. mai 2026 (i tide)", "–6 uker"),
        ("Vedtak", "—", "Kommunestyret sak NHB-2026-15", "Godkjent 07.05.2026"),
        ("Endringsdokument", "—", "Endringsdokument NHB-2026-15 - Schedule crashing.docx", "—"),
    ]
    for i, r in enumerate(rows, start=6):
        for c, val in enumerate(r, 1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.border = BORDER
            cell.alignment = WRAP

    set_col_widths(ws, [25, 35, 35, 30])


def build_tracking_per_måned(wb):
    """En rad per (måned, pakke) med % fullført og påløpt i perioden."""
    ws = wb.create_sheet("Tracking per måned")
    ws["A1"] = "Tracking-data per statusdato — verdier å skrive inn i MS Project"
    ws["A1"].font = Font(size=14, bold=True, color="1F4E78")
    ws.merge_cells("A1:G1")

    ws["A2"] = ("For hver måned: sett Project Status Date til 'Statusdato'-kolonnen, "
                "deretter for hver listede pakke fyll inn '% complete kum.' og 'Actual Cost kum.' "
                "(eller 'Actual Cost periode' om du foretrekker per-periode). "
                "Pakker som ikke er listet i en gitt måned er enten 0 % eller 100 % i den måneden.")
    ws["A2"].alignment = WRAP
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 50

    headers = ["Mnd", "Statusdato", "Pakke (WBS)", "Aktivitet", "% complete kum.",
               "Påløpt i perioden (MNOK)", "Akkumulert AC pakke (MNOK)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, len(headers))

    # Beregn kumulativ påløp per pakke
    kum_per_pakke: dict[str, float] = {p.wbs: 0.0 for p in PAKKER}
    r = 5
    for m in range(1, 17):
        navn, statusdato = MÅNEDER[m]
        # List bare pakker som er nevnt i referatet for denne måneden
        fremdrift_m = FREMDRIFT_PER_MND.get(m, {})
        # Sortér: pakker med eksplisitt data først, deretter _øvrige
        listet = [(w, data) for w, data in fremdrift_m.items() if w != "_øvrige"]
        for w, (pct, påløp) in listet:
            kum_per_pakke[w] += påløp
            p = PAKKE_BY_WBS[w]
            ws.cell(row=r, column=1, value=m).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=2, value=statusdato)
            ws.cell(row=r, column=3, value=w)
            ws.cell(row=r, column=4, value=p.navn)
            ws.cell(row=r, column=5, value=pct / 100).number_format = '0 %'
            ws.cell(row=r, column=6, value=påløp).number_format = '0.0'
            ws.cell(row=r, column=7, value=kum_per_pakke[w]).number_format = '0.0'
            for c in range(1, 8):
                ws.cell(row=r, column=c).border = BORDER
            r += 1
        # _øvrige som info-linje
        if "_øvrige" in fremdrift_m:
            øvrige = fremdrift_m["_øvrige"]
            ws.cell(row=r, column=1, value=m).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=2, value=statusdato)
            ws.cell(row=r, column=3, value="—")
            ws.cell(row=r, column=4, value="Øvrige aktive pakker (samlet)")
            ws.cell(row=r, column=5, value="")
            ws.cell(row=r, column=6, value=øvrige).number_format = '0.0'
            ws.cell(row=r, column=7, value="—")
            for c in range(1, 8):
                ws.cell(row=r, column=c).border = BORDER
                ws.cell(row=r, column=c).fill = ACCENT_FILL
            r += 1
        # Tom separator-rad
        r += 1

    set_col_widths(ws, [5, 14, 12, 36, 14, 18, 18])


def build_oppsummering(wb):
    """Forenklet oversikt per måned med PV/EV/AC fra master data."""
    ws = wb.create_sheet("Mnd-oversikt (PV-EV-AC)")
    ws["A1"] = "Måned-for-måned PV/EV/AC for direkte sammenligning i MS Project"
    ws["A1"].font = Font(size=14, bold=True, color="1F4E78")
    ws.merge_cells("A1:H1")

    headers = ["Mnd", "Periode", "Statusdato", "PV kum (MNOK)", "EV kum (MNOK)",
               "AC kum (MNOK)", "CPI", "SPI"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(headers))

    from log565_master_data import alle_måneder
    for m_data in alle_måneder():
        r = 3 + m_data.måned
        ws.cell(row=r, column=1, value=m_data.måned).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=m_data.måned_navn)
        ws.cell(row=r, column=3, value=m_data.statusdato)
        ws.cell(row=r, column=4, value=m_data.pv_kum).number_format = '0.0'
        ws.cell(row=r, column=5, value=m_data.ev_kum).number_format = '0.0'
        ws.cell(row=r, column=6, value=m_data.ac_kum).number_format = '0.0'
        ws.cell(row=r, column=7, value=m_data.cpi).number_format = '0.000'
        ws.cell(row=r, column=8, value=m_data.spi).number_format = '0.000'
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = BORDER

    set_col_widths(ws, [5, 18, 12, 14, 14, 14, 8, 8])


def main():
    wb = openpyxl.Workbook()
    build_instruks(wb)
    build_baseline1_tasks(wb)
    build_crashing_endring(wb)
    build_tracking_per_måned(wb)
    build_oppsummering(wb)
    wb.save(OUT)
    print(f"Lagret: {OUT.name}")


if __name__ == "__main__":
    main()
